import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import ctypes
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import List, Optional
from PIL import Image, ImageTk

# Hide console window immediately
try:
    ctypes.windll.kernel32.FreeConsole()
except:
    pass

DATA_FILE = os.path.join(os.path.dirname(__file__), "jsa_records.json")

@dataclass
class WorkStep:
    number: int
    work_step: str
    hazard: str
    control_method: str
    notes: str = ""

@dataclass
class JSARecord:
    id: int = 0
    job_title: str = ""
    workplace: str = ""
    date: str = ""
    analyst: str = ""
    freq_desc: str = ""
    freq_score: int = 1
    likelihood_desc: str = ""
    likelihood_score: int = 1
    severity_desc: str = ""
    severity_score: int = 1
    risk_score: int = 1
    risk_level: str = ""
    ppe_items: List[str] = field(default_factory=list)
    steps: List[WorkStep] = field(default_factory=list)
    worker: str = ""
    worker_sig: str = ""
    safety_officer: str = ""
    dept_manager: str = ""
    general_manager: str = ""
    department: str = ""
    general_notes: str = ""

FREQ_OPTIONS = [
    ("6 - بشكل مستمر (أو عدة مرات في اليوم)", 6),
    ("5 - بشكل متكرر - مرة واحدة يوميًا", 5),
    ("4 - أحيانًا - من مرة أسبوعيًا لشهريًا", 4),
    ("3 - عادة - من مرة شهريًا لسنويًا", 3),
    ("2 - نادرًا - كان من المعروف حدوثه", 2),
    ("1 - نادرًا جدًا - لا يُعرف حدوثه", 1),
]

LIKE_OPTIONS = [
    ("6 - النتيجة الأكثر احتمالا والمتوقعة", 6),
    ("5 - ممكن تماما - فرصة 50/50", 5),
    ("4 - تسلسل غير عادي أو مصادفة", 4),
    ("3 - ممكن عن بعد - من المعروف أن يحدث", 3),
    ("2 - بعيد جدًا ولكنه ممكن - لم يحدث أبدًا", 2),
    ("1 - مستحيل عمليا - واحد في المليون", 1),
]

SEV_OPTIONS = [
    ("6 - كارثية - العديد من الوفيات وأضرار جسيمة", 6),
    ("5 - عدة وفيات - 500k-1000k جنيه أضرار", 5),
    ("4 - وفاة - 100k-500k جنيه أضرار", 4),
    ("3 - إصابة خطيرة - 1000-100k جنيه أضرار", 3),
    ("2 - إصابات معيقة - حتى 1000 جنيه", 2),
    ("1 - جروح طفيفة وكدمات وأضرار طفيفة", 1),
]

RISK_MATRIX = {
    (1,1):1,(1,2):2,(1,3):3,(1,4):4,(1,5):5,(1,6):6,
    (2,1):2,(2,2):4,(2,3):6,(2,4):8,(2,5):10,(2,6):12,
    (3,1):3,(3,2):6,(3,3):9,(3,4):12,(3,5):15,(3,6):18,
    (4,1):4,(4,2):8,(4,3):12,(4,4):16,(4,5):20,(4,6):24,
    (5,1):5,(5,2):10,(5,3):15,(5,4):20,(5,5):25,(5,6):30,
    (6,1):6,(6,2):12,(6,3):18,(6,4):24,(6,5):30,(6,6):36,
}

def get_risk_level(score):
    if score <= 5:
        return "عادية", "#4CAF50"
    elif score <= 15:
        return "بسيطة", "#8BC34A"
    elif score <= 25:
        return "متوسطة", "#FFC107"
    elif score <= 50:
        return "خطرة", "#FF9800"
    elif score <= 100:
        return "شديدة الخطورة", "#F44336"
    else:
        return "كارثية", "#B71C1C"

class JSADatabase:
    DEFAULT_DEPARTMENTS = ["الهندسية", "الجودة", "اللوجيستك", "الأمن", "الإنتاج",
                           "الشئون الإدارية", "المخازن", "المشتريات", "السلامة",
                           "الصيانة", "الموارد البشرية", "نظم المعلومات", "الحسابات",
                           "المراجعة", "الإدارة العليا", "أخرى"]
    DEFAULT_JOB_TITLES = []

    def __init__(self):
        self.records = []
        self.next_id = 1
        self.departments = list(self.DEFAULT_DEPARTMENTS)
        self.job_titles = list(self.DEFAULT_JOB_TITLES)
        self.load()

    def load(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.records = [self._dict_to_record(r) for r in data.get("records", [])]
                self.next_id = data.get("next_id", max([r.id for r in self.records], default=0) + 1)
                self.departments = data.get("departments", list(self.DEFAULT_DEPARTMENTS))
                self.job_titles = data.get("job_titles", list(self.DEFAULT_JOB_TITLES))
            except:
                self.records = []
                self.next_id = 1
                self.departments = list(self.DEFAULT_DEPARTMENTS)
                self.job_titles = list(self.DEFAULT_JOB_TITLES)

    def save(self):
        os.makedirs(os.path.dirname(DATA_FILE) or ".", exist_ok=True)
        data = {
            "records": [asdict(r) for r in self.records],
            "next_id": self.next_id,
            "departments": self.departments,
            "job_titles": self.job_titles,
        }
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _dict_to_record(self, d):
        steps = [WorkStep(**s) for s in d.get("steps", [])]
        d["steps"] = steps
        return JSARecord(**d)

    def add(self, record: JSARecord):
        record.id = self.next_id
        self.next_id += 1
        self.records.append(record)
        self.save()

    def update(self, record: JSARecord):
        for i, r in enumerate(self.records):
            if r.id == record.id:
                self.records[i] = record
                self.save()
                return

    def delete(self, record_id: int):
        self.records = [r for r in self.records if r.id != record_id]
        self.save()

    def search(self, query: str = ""):
        if not query.strip():
            return self.records
        q = query.strip().lower()
        results = []
        for r in self.records:
            if (q in r.job_title.lower() or q in r.workplace.lower() or
                q in r.analyst.lower() or q in r.risk_level.lower() or
                q in r.department.lower() or
                any(q in s.hazard.lower() or q in s.work_step.lower() for s in r.steps) or
                any(q in p.lower() for p in r.ppe_items)):
                results.append(r)
        return results

db = JSADatabase()

class JSAListView(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg="#f0f4f8")
        self.app = app
        self.build_ui()

    def build_ui(self):
        top = tk.Frame(self, bg="#1a237e")
        top.pack(fill="x")

        logo_path = os.path.join(os.path.dirname(__file__), "namaa logo.png")
        for p in [logo_path, logo_path.replace("namaa logo.png", "nama_logo.png")]:
            if os.path.exists(p):
                logo_path = p
                break
        try:
            logo_img = ImageTk.PhotoImage(Image.open(logo_path).resize((80, 80), Image.LANCZOS))
            logo_lbl = tk.Label(top, image=logo_img, bg="#1a237e")
            logo_lbl.image = logo_img
            logo_lbl.pack(side="right", padx=(10, 5), pady=5)
        except:
            pass

        texts = tk.Frame(top, bg="#1a237e")
        texts.pack(side="right", padx=(5, 20))
        tk.Label(texts, text="شركة نماء لصناعة الأعلاف", font=("Arial", 18, "bold"),
                 bg="#1a237e", fg="white").pack(anchor="e")
        tk.Label(texts, text="نظام تحليل مخاطر الوظيفة (JSA)", font=("Arial", 13),
                 bg="#1a237e", fg="#B0BEC5").pack(anchor="e")
        tk.Label(texts, text="by Eng. Bahaa Mohamed", font=("Arial", 10),
                 bg="#1a237e", fg="#90A4AE").pack(anchor="e")

        search_frame = tk.Frame(self, bg="#f0f4f8")
        search_frame.pack(fill="x", padx=15, pady=10)
        tk.Label(search_frame, text="بحث:", font=("Arial", 11), bg="#f0f4f8").pack(side="right", padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *a: self.search())
        tk.Entry(search_frame, textvariable=self.search_var, font=("Arial", 11), width=40).pack(side="right", padx=5)

        DEPARTMENTS = ["الكل"] + db.departments
        self.dept_var = tk.StringVar(value="الكل")
        tk.Label(search_frame, text="الإدارة:", font=("Arial", 11), bg="#f0f4f8").pack(side="right", padx=(15, 5))
        self.dept_menu = ttk.Combobox(search_frame, textvariable=self.dept_var, values=DEPARTMENTS,
                                      width=18, state="readonly", font=("Arial", 9))
        self.dept_menu.pack(side="right", padx=5)
        self.dept_var.trace_add("write", lambda *a: self.search())

        btn_frame = tk.Frame(search_frame, bg="#f0f4f8")
        btn_frame.pack(side="left")
        tk.Button(btn_frame, text="إضافة JSA جديد", font=("Arial", 10, "bold"),
                  bg="#4CAF50", fg="white", padx=12, pady=4,
                  command=self.app.new_record).pack(side="left", padx=3)
        tk.Button(btn_frame, text="تقرير PDF", font=("Arial", 10, "bold"),
                  bg="#FF9800", fg="white", padx=12, pady=4,
                  command=self.app.generate_report).pack(side="left", padx=3)

        cols = ("مستوى الخطورة", "درجة المخاطرة", "المحلل", "مكان العمل", "الإدارة", "الوظيفة", "#")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18, selectmode="browse")
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor="center")
        self.tree.column("#", width=40, anchor="center")
        self.tree.column("الوظيفة", width=220, anchor="e")
        self.tree.column("الإدارة", width=140, anchor="e")
        self.tree.column("مكان العمل", width=180, anchor="e")
        self.tree.column("المحلل", width=140, anchor="e")
        self.tree.column("درجة المخاطرة", width=80, anchor="center")
        self.tree.column("مستوى الخطورة", width=100, anchor="center")

        self._build_btns()

        scroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="right", fill="both", expand=True, padx=(0,15), pady=(0,15))
        scroll.pack(side="left", fill="y", pady=(0,15))

        self.tree.bind("<Double-1>", lambda e: self.app.view_record())
        self.tree.bind("<Return>", lambda e: self.app.view_record())
        self.tree.bind("<MouseWheel>", self._on_mw)
        self.tree.bind("<Button-4>", lambda e: self.tree.yview_scroll(-3, "units"))
        self.tree.bind("<Button-5>", lambda e: self.tree.yview_scroll(3, "units"))
        self.app.root.bind("<MouseWheel>", self._on_mw_root, add="+")

    def _on_mw(self, event):
        self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
    def _on_mw_root(self, event):
        x, y = self.winfo_pointerxy()
        w = self.winfo_containing(x, y)
        if w and w.winfo_toplevel() == self.winfo_toplevel():
            self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _build_btns(self):
        btn_frame2 = tk.Frame(self, bg="#f0f4f8")
        btn_frame2.pack(fill="x", padx=15, pady=(0,10))
        tk.Button(btn_frame2, text="فتح", font=("Arial", 10), bg="#2196F3", fg="white",
                  padx=20, command=self.app.view_record).pack(side="left", padx=3)
        tk.Button(btn_frame2, text="تعديل", font=("Arial", 10), bg="#FF9800", fg="white",
                  padx=20, command=self.app.edit_record).pack(side="left", padx=3)
        tk.Button(btn_frame2, text="حذف", font=("Arial", 10), bg="#f44336", fg="white",
                  padx=20, command=self.app.delete_record).pack(side="left", padx=3)
        tk.Button(btn_frame2, text="تحليلات", font=("Arial", 10, "bold"),
                  bg="#9C27B0", fg="white", padx=15,
                  command=self.app.open_analysis).pack(side="left", padx=3)
        tk.Button(btn_frame2, text="إدارة البيانات", font=("Arial", 10, "bold"),
                  bg="#607D8B", fg="white", padx=15,
                  command=self.app.open_dept_mgr).pack(side="left", padx=3)
        tk.Button(btn_frame2, text="استيراد من إكسيل", font=("Arial", 10, "bold"),
                  bg="#009688", fg="white", padx=15,
                  command=self.app.import_excel).pack(side="left", padx=3)

        self.refresh()

    def search(self):
        self.refresh()

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        results = db.search(self.search_var.get())
        dept = self.dept_var.get()
        if dept and dept != "الكل":
            results = [r for r in results if r.department == dept]
        for r in results:
            level, color = get_risk_level(r.risk_score)
            self.tree.insert("", "end", values=(
                level, r.risk_score, r.analyst, r.workplace, r.department, r.job_title, r.id
            ))

    def get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("تنبيه", "الرجاء تحديد سجل")
            return None
        vals = self.tree.item(sel[0], "values")
        return int(vals[-1])  # id is the last value

    def _update_dept_filter(self):
        vals = ["الكل"] + db.departments
        self.dept_menu["values"] = vals
        if self.dept_var.get() not in vals:
            self.dept_var.set("الكل")

class CollapsibleSection(tk.Frame):
    def __init__(self, parent, title, bg="#f0f4f8", expanded=True):
        super().__init__(parent, bg=bg)
        self.expanded = expanded
        self._bg = bg

        hdr = tk.Frame(self, bg="#e8eaf6", cursor="hand2", height=35)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        hdr.grid_columnconfigure(1, weight=1)

        self.btn = tk.Label(hdr, text="▲" if expanded else "▼",
                            font=("Arial", 10, "bold"),
                            bg="#e8eaf6", fg="#1a237e")
        self.btn.grid(row=0, column=0, padx=(15, 5), sticky="w")

        title_lbl = tk.Label(hdr, text=title, font=("Arial", 12, "bold"),
                             bg="#e8eaf6", fg="#1a237e")
        title_lbl.grid(row=0, column=1, sticky="ew", padx=5)

        spacer = tk.Frame(hdr, bg="#e8eaf6", width=50)
        spacer.grid(row=0, column=2, padx=(5, 15))

        for w in (hdr, self.btn, title_lbl, spacer):
            w.bind("<Button-1>", self.toggle)

        self.content = tk.Frame(self, bg=bg)
        if expanded:
            self.content.pack(fill="x", padx=15, pady=15)

    def toggle(self, event=None):
        if self.expanded:
            self.content.pack_forget()
            self.btn.config(text="▼")
        else:
            self.content.pack(fill="x", padx=15, pady=15)
            self.btn.config(text="▲")
        self.expanded = not self.expanded


class DeptManager(tk.Toplevel):
    def __init__(self, app):
        super().__init__(app.root)
        self.app = app
        self.title("إدارة البيانات")
        self.geometry("650x450")
        self.configure(bg="#f0f4f8")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=10)

        # ----- Departments tab -----
        dept_frame = tk.Frame(nb, bg="#f0f4f8")
        nb.add(dept_frame, text="الإدارات")

        tk.Label(dept_frame, text="قائمة الإدارات", font=("Arial", 12, "bold"),
                 bg="#1a237e", fg="white").pack(fill="x", padx=5, pady=5)

        list_frame = tk.Frame(dept_frame, bg="#f0f4f8")
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.dept_listbox = tk.Listbox(list_frame, font=("Arial", 11))
        scroll_d = ttk.Scrollbar(list_frame, orient="vertical", command=self.dept_listbox.yview)
        self.dept_listbox.configure(yscrollcommand=scroll_d.set)
        self.dept_listbox.pack(side="right", fill="both", expand=True)
        scroll_d.pack(side="left", fill="y")
        self._fill_depts()

        btn_d = tk.Frame(dept_frame, bg="#f0f4f8")
        btn_d.pack(fill="x", padx=5, pady=5)
        self.dept_entry = tk.Entry(btn_d, font=("Arial", 11), width=30)
        self.dept_entry.pack(side="right", padx=5)
        tk.Button(btn_d, text="إضافة", font=("Arial", 10, "bold"), bg="#4CAF50",
                  fg="white", command=self._add_dept).pack(side="right", padx=2)
        tk.Button(btn_d, text="حذف المحدد", font=("Arial", 10, "bold"), bg="#f44336",
                  fg="white", command=self._del_dept).pack(side="right", padx=2)

        # ----- Job Titles tab -----
        job_frame = tk.Frame(nb, bg="#f0f4f8")
        nb.add(job_frame, text="الوظائف")

        tk.Label(job_frame, text="قائمة الوظائف", font=("Arial", 12, "bold"),
                 bg="#1a237e", fg="white").pack(fill="x", padx=5, pady=5)

        list_frame_j = tk.Frame(job_frame, bg="#f0f4f8")
        list_frame_j.pack(fill="both", expand=True, padx=5, pady=5)
        self.job_listbox = tk.Listbox(list_frame_j, font=("Arial", 11))
        scroll_j = ttk.Scrollbar(list_frame_j, orient="vertical", command=self.job_listbox.yview)
        self.job_listbox.configure(yscrollcommand=scroll_j.set)
        self.job_listbox.pack(side="right", fill="both", expand=True)
        scroll_j.pack(side="left", fill="y")
        self._fill_jobs()

        btn_j = tk.Frame(job_frame, bg="#f0f4f8")
        btn_j.pack(fill="x", padx=5, pady=5)
        self.job_entry = tk.Entry(btn_j, font=("Arial", 11), width=30)
        self.job_entry.pack(side="right", padx=5)
        tk.Button(btn_j, text="إضافة", font=("Arial", 10, "bold"), bg="#4CAF50",
                  fg="white", command=self._add_job).pack(side="right", padx=2)
        tk.Button(btn_j, text="حذف المحدد", font=("Arial", 10, "bold"), bg="#f44336",
                  fg="white", command=self._del_job).pack(side="right", padx=2)

        tk.Button(self, text="إغلاق", font=("Arial", 11), bg="#f44336", fg="white",
                  padx=30, command=self.destroy).pack(pady=10)

    def _fill_depts(self):
        self.dept_listbox.delete(0, "end")
        for d in db.departments:
            self.dept_listbox.insert("end", d)
    def _fill_jobs(self):
        self.job_listbox.delete(0, "end")
        for j in db.job_titles:
            self.job_listbox.insert("end", j)
    def _add_dept(self):
        v = self.dept_entry.get().strip()
        if v and v not in db.departments:
            db.departments.append(v)
            db.save()
            self._fill_depts()
            self.dept_entry.delete(0, "end")
            self.app._refresh_depts()
    def _del_dept(self):
        sel = self.dept_listbox.curselection()
        if sel:
            v = self.dept_listbox.get(sel[0])
            db.departments.remove(v)
            db.save()
            self._fill_depts()
            self.app._refresh_depts()
    def _add_job(self):
        v = self.job_entry.get().strip()
        if v and v not in db.job_titles:
            db.job_titles.append(v)
            db.save()
            self._fill_jobs()
            self.job_entry.delete(0, "end")
    def _del_job(self):
        sel = self.job_listbox.curselection()
        if sel:
            v = self.job_listbox.get(sel[0])
            db.job_titles.remove(v)
            db.save()
            self._fill_jobs()


class AnalysisDashboard(tk.Toplevel):
    def __init__(self, app):
        super().__init__(app.root)
        self.app = app
        self.title("تحليلات المخاطر")
        self.geometry("950x700")
        self.configure(bg="white")

        canvas = tk.Canvas(self, bg="white")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        frame = tk.Frame(canvas, bg="white")
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw", width=920)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        def _mw(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _mw)
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _mw))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        records = db.records

        # -- Summary --
        sum_f = tk.LabelFrame(frame, text="ملخص عام", font=("Arial", 12, "bold"),
                              bg="white", fg="#1a237e", padx=15, pady=10)
        sum_f.pack(fill="x", padx=20, pady=10)

        total = len(records)
        risk_levels = ["عادية", "بسيطة", "متوسطة", "خطرة", "شديدة الخطورة", "كارثية"]
        level_counts = {l: 0 for l in risk_levels}
        dept_counts = {}
        for r in records:
            lvl, _ = get_risk_level(r.risk_score)
            level_counts[lvl] = level_counts.get(lvl, 0) + 1
            dept_counts[r.department] = dept_counts.get(r.department, 0) + 1

        for i, (lbl, val) in enumerate([
            ("إجمالي السجلات:", str(total)),
            ("عدد الإدارات:", str(len(dept_counts))),
            ("الوظائف المسجلة:", str(len(set(r.job_title for r in records if r.job_title)))),
        ]):
            row = tk.Frame(sum_f, bg="white")
            row.pack(fill="x", pady=2)
            tk.Label(row, text=lbl, font=("Arial", 10, "bold"), bg="white",
                     width=20, anchor="e").pack(side="right")
            tk.Label(row, text=val, font=("Arial", 10), bg="#f5f5f5",
                     anchor="w", padx=5).pack(side="right", fill="x", expand=True)

        # -- Risk Level Distribution (bar chart) --
        risk_f = tk.LabelFrame(frame, text="توزيع مستويات الخطورة", font=("Arial", 12, "bold"),
                               bg="white", fg="#1a237e", padx=15, pady=10)
        risk_f.pack(fill="x", padx=20, pady=10)

        colors_map = {"عادية": "#4CAF50", "بسيطة": "#8BC34A", "متوسطة": "#FFC107",
                      "خطرة": "#FF9800", "شديدة الخطورة": "#F44336", "كارثية": "#B71C1C"}
        max_cnt = max(level_counts.values()) or 1
        bar_max = 400
        for lvl in risk_levels:
            cnt = level_counts.get(lvl, 0)
            row = tk.Frame(risk_f, bg="white")
            row.pack(fill="x", pady=3)
            bar_w = int((cnt / max_cnt) * bar_max) if cnt else 0
            tk.Label(row, text=f"{lvl}: {cnt}", font=("Arial", 9), bg="white",
                     width=20, anchor="e").pack(side="right")
            if bar_w:
                bar = tk.Frame(row, bg=colors_map.get(lvl, "#999"),
                               width=bar_w, height=18)
                bar.pack(side="right", padx=2)
                bar.pack_propagate(False)

        # -- Department Analysis --
        dept_f = tk.LabelFrame(frame, text="تحليل الإدارات", font=("Arial", 12, "bold"),
                               bg="white", fg="#1a237e", padx=15, pady=10)
        dept_f.pack(fill="x", padx=20, pady=10)

        dept_sorted = sorted(dept_counts.items(), key=lambda x: -x[1])
        for dept, cnt in dept_sorted:
            if not dept:
                continue
            records_dept = [r for r in records if r.department == dept]
            avg_score = sum(r.risk_score for r in records_dept) / max(cnt, 1)
            high_risk = sum(1 for r in records_dept if r.risk_score > 25)
            dept_lvl, _ = get_risk_level(int(avg_score))

            sec = tk.Frame(dept_f, bg="white", relief="solid", bd=1)
            sec.pack(fill="x", pady=3, padx=2)
            inner = tk.Frame(sec, bg="#fafafa")
            inner.pack(fill="x", padx=5, pady=4)
            for i, (lbl, val) in enumerate([
                ("الإدارة:", dept),
                ("عدد السجلات:", str(cnt)),
                ("متوسط الخطورة:", f"{avg_score:.1f} ({dept_lvl})"),
                ("سجلات عالية الخطورة:", str(high_risk)),
            ]):
                tk.Label(inner, text=lbl, font=("Arial", 9, "bold" if i == 0 else "normal"),
                         bg="#fafafa", width=18, anchor="e").pack(side="right")
                tk.Label(inner, text=val, font=("Arial", 9),
                         bg="#fafafa", anchor="w").pack(side="right", padx=5)

        # -- High Risk Records --
        high_f = tk.LabelFrame(frame, text="السجلات عالية الخطورة", font=("Arial", 12, "bold"),
                               bg="white", fg="#1a237e", padx=15, pady=10)
        high_f.pack(fill="x", padx=20, pady=10)

        high_records = [r for r in records if r.risk_score > 25]
        if high_records:
            for r in high_records[:10]:
                lvl, col = get_risk_level(r.risk_score)
                row = tk.Frame(high_f, bg="white")
                row.pack(fill="x", pady=1)
                tk.Label(row, text=f"{r.job_title} | {r.department} | درجة {r.risk_score} | {lvl}",
                         font=("Arial", 9), bg="white", anchor="w",
                         fg=col).pack(side="right", padx=5)
        else:
            tk.Label(high_f, text="لا توجد سجلات عالية الخطورة", font=("Arial", 10),
                     bg="white", anchor="w").pack(fill="x", padx=5, pady=5)

        # -- Export button --
        btn_f = tk.Frame(frame, bg="white")
        btn_f.pack(pady=15)
        tk.Button(btn_f, text="طباعة التحليل PDF", font=("Arial", 11, "bold"),
                  bg="#FF9800", fg="white", padx=25, pady=6,
                  command=self.export_pdf).pack(side="right", padx=5)
        tk.Button(btn_f, text="إغلاق", font=("Arial", 11), bg="#f44336",
                  fg="white", padx=25, command=self.destroy).pack(side="right", padx=5)

    def export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as pdf_canvas
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import arabic_reshaper
            from bidi.algorithm import get_display

            def ar(t):
                return get_display(arabic_reshaper.reshape(t))

            pdfmetrics.registerFont(TTFont('Ar', r"C:\Windows\Fonts\arial.ttf"))
            pdfmetrics.registerFont(TTFont('ArBd', r"C:\Windows\Fonts\arialbd.ttf"))

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile="تحليل_المخاطر.pdf"
            )
            if not file_path:
                return

            pdf = pdf_canvas.Canvas(file_path, pagesize=A4)
            W, H = A4
            RM = W - 30
            y = H - 40

            pdf.setFont("ArBd", 18)
            pdf.drawRightString(RM, y, ar("تحليل المخاطر - شركة نماء"))
            y -= 25
            pdf.setFont("Ar", 10)
            from datetime import date
            pdf.drawRightString(RM, y, ar(f"تاريخ التقرير: {date.today()}"))
            y -= 30

            records = db.records
            risk_levels = ["عادية", "بسيطة", "متوسطة", "خطرة", "شديدة الخطورة", "كارثية"]
            level_counts = {l: 0 for l in risk_levels}
            dept_data = {}
            for r in records:
                lvl, _ = get_risk_level(r.risk_score)
                level_counts[lvl] = level_counts.get(lvl, 0) + 1
                if r.department not in dept_data:
                    dept_data[r.department] = {"count": 0, "total_score": 0, "high": 0}
                dept_data[r.department]["count"] += 1
                dept_data[r.department]["total_score"] += r.risk_score
                if r.risk_score > 25:
                    dept_data[r.department]["high"] += 1

            # Summary section
            def check_page(needed=60):
                nonlocal y
                if y < needed:
                    pdf.showPage()
                    y = H - 40

            def sec_title(txt):
                nonlocal y
                check_page(70)
                pdf.setFillColorRGB(0.91, 0.92, 0.96)
                pdf.rect(20, y - 2, RM - 20 + 20, 22, fill=1, stroke=0)
                pdf.setFillColor(colors.black)
                pdf.setFont("ArBd", 12)
                pdf.drawRightString(RM, y + 4, ar(txt))
                y -= 35

            sec_title("ملخص عام")
            pdf.setFont("Ar", 11)
            for lbl, val in [("إجمالي السجلات:", len(records)),
                             ("عدد الإدارات:", len(dept_data))]:
                pdf.drawRightString(RM, y, ar(f"{lbl} {val}"))
                y -= 28

            sec_title("توزيع مستويات الخطورة")
            pdf.setFont("Ar", 11)
            for lvl in risk_levels:
                cnt = level_counts.get(lvl, 0)
                pdf.drawRightString(RM, y, ar(f"{lvl}: {cnt}"))
                y -= 24

            sec_title("تحليل الإدارات")
            pdf.setFont("Ar", 11)
            for dept, d in sorted(dept_data.items(), key=lambda x: -x[1]["count"]):
                if not dept:
                    continue
                avg = d["total_score"] / d["count"]
                dlvl, _ = get_risk_level(int(avg))
                check_page(60)
                pdf.drawRightString(RM, y, ar(f"الإدارة: {dept}"))
                y -= 24
                check_page(60)
                pdf.setFont("Ar", 10)
                pdf.drawRightString(RM, y, ar(f"عدد السجلات: {d['count']} | "
                                              f"متوسط الخطورة: {avg:.0f} ({dlvl}) | "
                                              f"عالية الخطورة: {d['high']}"))
                pdf.setFont("Ar", 11)
                y -= 30

            pdf.save()
            messagebox.showinfo("نجاح", f"تم حفظ التقرير:\n{file_path}")
        except Exception as e:
            import traceback
            messagebox.showerror("خطأ", f"فشل في إنشاء التقرير:\n{str(e)}\n\n{traceback.format_exc()}")


class JSAFormView(tk.Toplevel):
    def __init__(self, app, record: Optional[JSARecord] = None):
        super().__init__(app.root)
        self.app = app
        self.record = record or JSARecord()
        self.is_new = record is None
        title = "إضافة JSA جديد" if self.is_new else f"تعديل JSA - {record.job_title}"
        self.title(title)
        self.geometry("1100x750")
        self.configure(bg="#f0f4f8")
        self.build_ui()
        if not self.is_new:
            self.load_record()

    def build_ui(self):
        canvas = tk.Canvas(self, bg="#f0f4f8")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#f0f4f8")

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw", width=1060)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        self._scroll_binding = _on_mousewheel

        def _on_enter(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Enter>", _on_enter)

        def _on_leave(event):
            canvas.unbind_all("<MouseWheel>")
        canvas.bind("<Leave>", _on_leave)

        row = -1
        top_form = tk.Frame(scroll_frame, bg="#1a237e")
        top_form.grid(row=0, column=0, columnspan=2, sticky="ew", padx=0, pady=0)
        logo_path = os.path.join(os.path.dirname(__file__), "namaa logo.png")
        for p in [logo_path, logo_path.replace("namaa logo.png", "nama_logo.png")]:
            if os.path.exists(p):
                logo_path = p
                break
        try:
            img = ImageTk.PhotoImage(Image.open(logo_path).resize((70, 70), Image.LANCZOS))
            lbl = tk.Label(top_form, image=img, bg="#1a237e")
            lbl.image = img
            lbl.grid(row=0, column=0, padx=(10, 5), pady=5)
        except:
            pass
        text_f = tk.Frame(top_form, bg="#1a237e")
        text_f.grid(row=0, column=1, sticky="w", padx=(5, 10))
        tk.Label(text_f, text="شركة نماء لصناعة الأعلاف", font=("Arial", 16, "bold"),
                 bg="#1a237e", fg="white").pack(anchor="e")
        top_form.columnconfigure(1, weight=1)

        row = 1
        sec_basic = CollapsibleSection(scroll_frame, "البيانات الأساسية")
        sec_basic.grid(row=row, column=0, padx=20, pady=(5, 10), sticky="ew", columnspan=2)
        sec = sec_basic.content
        sec.columnconfigure(1, weight=1)

        DEPARTMENTS = db.departments
        self.dept_form_var = tk.StringVar()
        fields = [
            ("JOB TITLE", "job_title", "الوظيفة:", "entry"),
            ("WORKPLACE", "workplace", "مكان العمل:", "entry"),
            ("DEPARTMENT", "department", "الإدارة:", "combo"),
            ("DATE", "date", "التاريخ:", "entry"),
            ("ANALYST", "analyst", "القائم بالتحليل:", "entry"),
            ("WORKER", "worker", "القائم بالعمل:", "entry"),
            ("GENERAL NOTES", "general_notes", "الملاحظات العامة:", "text"),
        ]
        for i, (eng, key, label, etype) in enumerate(fields):
            enlbl = tk.Label(sec, text=eng, font=("Arial", 9, "bold"), bg="#f0f4f8", fg="#555")
            enlbl.grid(row=i, column=0, sticky="e", padx=(5, 10), pady=3)
            if etype == "text":
                ent = tk.Text(sec, height=3, font=("Arial", 10))
            elif etype == "combo":
                ent = ttk.Combobox(sec, textvariable=self.dept_form_var, values=DEPARTMENTS,
                                   state="readonly", font=("Arial", 10))
            else:
                ent = tk.Entry(sec, font=("Arial", 10))
            ent.grid(row=i, column=1, sticky="ew", padx=5, pady=3)
            lbl = tk.Label(sec, text=label, font=("Arial", 10, "bold"), bg="#f0f4f8")
            lbl.grid(row=i, column=2, sticky="w", padx=(10, 5), pady=3)
            setattr(self, key, ent)

        row += 1
        sec_risk = CollapsibleSection(scroll_frame, "تقييم المخاطر")
        sec_risk.grid(row=row, column=0, padx=20, pady=(5, 10), sticky="ew", columnspan=2)
        risk_frame = sec_risk.content
        risk_frame.columnconfigure(1, weight=1)

        self.freq_var = tk.StringVar()
        self.like_var = tk.StringVar()
        self.sev_var = tk.StringVar()

        tk.Label(risk_frame, text="FREQUENCY", font=("Arial", 9, "bold"), bg="#f0f4f8", fg="#555").grid(row=0, column=0, sticky="e", padx=(5, 10), pady=3)
        freq_menu = ttk.Combobox(risk_frame, textvariable=self.freq_var, values=[o[0] for o in FREQ_OPTIONS],
                                 state="readonly", font=("Arial", 9))
        freq_menu.grid(row=0, column=1, sticky="ew", padx=5, pady=3)
        tk.Label(risk_frame, text="التكرار:", font=("Arial", 10, "bold"), bg="#f0f4f8").grid(row=0, column=2, sticky="w", padx=(10, 5), pady=3)

        tk.Label(risk_frame, text="LIKELIHOOD", font=("Arial", 9, "bold"), bg="#f0f4f8", fg="#555").grid(row=1, column=0, sticky="e", padx=(5, 10), pady=3)
        like_menu = ttk.Combobox(risk_frame, textvariable=self.like_var, values=[o[0] for o in LIKE_OPTIONS],
                                 state="readonly", font=("Arial", 9))
        like_menu.grid(row=1, column=1, sticky="ew", padx=5, pady=3)
        tk.Label(risk_frame, text="الاحتمالية:", font=("Arial", 10, "bold"), bg="#f0f4f8").grid(row=1, column=2, sticky="w", padx=(10, 5), pady=3)

        tk.Label(risk_frame, text="SEVERITY", font=("Arial", 9, "bold"), bg="#f0f4f8", fg="#555").grid(row=2, column=0, sticky="e", padx=(5, 10), pady=3)
        sev_menu = ttk.Combobox(risk_frame, textvariable=self.sev_var, values=[o[0] for o in SEV_OPTIONS],
                                state="readonly", font=("Arial", 9))
        sev_menu.grid(row=2, column=1, sticky="ew", padx=5, pady=3)
        tk.Label(risk_frame, text="الخطورة:", font=("Arial", 10, "bold"), bg="#f0f4f8").grid(row=2, column=2, sticky="w", padx=(10, 5), pady=3)

        self.risk_score_label = tk.Label(risk_frame, text="درجة المخاطرة: --", font=("Arial", 11, "bold"),
                                         bg="#f0f4f8", fg="#1a237e")
        self.risk_score_label.grid(row=3, column=0, columnspan=3, pady=5)

        self.risk_level_label = tk.Label(risk_frame, text="مستوى الخطورة: --", font=("Arial", 11, "bold"),
                                         bg="#f0f4f8", fg="#1a237e")
        self.risk_level_label.grid(row=4, column=0, columnspan=3, pady=5)

        for var in (self.freq_var, self.like_var, self.sev_var):
            var.trace_add("write", lambda *a: self.calc_risk())

        row += 1
        sec_ppe = CollapsibleSection(scroll_frame, "مهمات الوقاية الشخصية (PPE)")
        sec_ppe.grid(row=row, column=0, padx=20, pady=(5, 10), sticky="ew", columnspan=2)
        ppe_frame = sec_ppe.content

        ppe_list = ["النظارات الواقية", "جوانتى", "درع لحام", "ملابس واقية", "أحذية للانزلاق",
                    "أحذية آمنة", "جهاز تنفس", "خوذة", "واقي سمع", "حزام أمان"]
        self.ppe_vars = {}
        inner = tk.Frame(ppe_frame, bg="#f0f4f8")
        inner.pack()
        for i, ppe in enumerate(ppe_list):
            var = tk.BooleanVar()
            cb = tk.Checkbutton(inner, text=ppe, variable=var, font=("Arial", 10), bg="#f0f4f8")
            cb.grid(row=i//5, column=i%5, padx=10, pady=2)
            self.ppe_vars[ppe] = var

        row += 1
        sec_steps = CollapsibleSection(scroll_frame, "خطوات العمل وتحليل المخاطر")
        sec_steps.grid(row=row, column=0, padx=20, pady=(5, 10), sticky="ew", columnspan=2)
        steps_frame = sec_steps.content

        cols = ["الملاحظات", "طرق التحكم", "الاخطار المحتمله", "خطوات العمل", "م"]
        self.step_widgets = []
        for c, col in enumerate(cols):
            tk.Label(steps_frame, text=col, font=("Arial", 9, "bold"), bg="#e8eaf6",
                     borderwidth=1, relief="solid", width=8 if c == 4 else 20).grid(row=0, column=c, padx=1, pady=2, sticky="ew")
        steps_frame.grid_columnconfigure(0, weight=1)
        steps_frame.grid_columnconfigure(1, weight=1)
        steps_frame.grid_columnconfigure(2, weight=1)
        steps_frame.grid_columnconfigure(3, weight=1)

        for i in range(10):
            row_num = i + 1
            ent_notes = tk.Text(steps_frame, height=2, width=15, font=("Arial", 9))
            ent_notes.grid(row=row_num, column=0, padx=1, pady=2, sticky="ew")
            ent_control = tk.Text(steps_frame, height=2, width=25, font=("Arial", 9))
            ent_control.grid(row=row_num, column=1, padx=1, pady=2, sticky="ew")
            ent_hazard = tk.Text(steps_frame, height=2, width=25, font=("Arial", 9))
            ent_hazard.grid(row=row_num, column=2, padx=1, pady=2, sticky="ew")
            ent_step = tk.Text(steps_frame, height=2, width=25, font=("Arial", 9))
            ent_step.grid(row=row_num, column=3, padx=1, pady=2, sticky="ew")
            ent_num = tk.Entry(steps_frame, width=6, font=("Arial", 10), justify="center")
            ent_num.grid(row=row_num, column=4, padx=1, pady=2)
            ent_num.insert(0, str(row_num))
            self.step_widgets.append((ent_num, ent_step, ent_hazard, ent_control, ent_notes))

        row += 1
        sec_sig = CollapsibleSection(scroll_frame, "التوقيعات")
        sec_sig.grid(row=row, column=0, padx=20, pady=(5, 10), sticky="ew", columnspan=2)
        sig_frame = sec_sig.content
        sig_frame.columnconfigure(1, weight=1)

        sig_data = [
            ("SAFETY OFFICER", "safety_officer", "السلامة والصحة المهنية:"),
            ("WORKER", "worker_sig", "القائم بالعملية:"),
            ("DEPT. MANAGER", "dept_manager", "مدير الإدارة:"),
            ("GENERAL MANAGER", "general_manager", "المدير العام:"),
        ]
        self.worker_sig = tk.Entry(sig_frame, font=("Arial", 10))
        for i, (eng, key, label) in enumerate(sig_data):
            enlbl = tk.Label(sig_frame, text=eng, font=("Arial", 9, "bold"), bg="#f0f4f8", fg="#555")
            enlbl.grid(row=i, column=0, sticky="e", padx=(5, 10), pady=3)
            if key == "worker_sig":
                ent = self.worker_sig
            else:
                ent = tk.Entry(sig_frame, font=("Arial", 10))
            ent.grid(row=i, column=1, sticky="ew", padx=5, pady=3)
            lbl = tk.Label(sig_frame, text=label, font=("Arial", 10, "bold"), bg="#f0f4f8")
            lbl.grid(row=i, column=2, sticky="w", padx=(10, 5), pady=3)
            setattr(self, key, ent)

        row += 1
        btn_save_frame = tk.Frame(scroll_frame, bg="#f0f4f8")
        btn_save_frame.grid(row=row, column=0, columnspan=2, pady=20)
        tk.Button(btn_save_frame, text="حفظ", font=("Arial", 12, "bold"), bg="#4CAF50",
                  fg="white", padx=30, command=self.save).pack(side="left", padx=10)
        tk.Button(btn_save_frame, text="إلغاء", font=("Arial", 12, "bold"), bg="#f44336",
                  fg="white", padx=30, command=self.destroy).pack(side="left", padx=10)

    def calc_risk(self):
        freq = self.freq_var.get()
        like = self.like_var.get()
        sev = self.sev_var.get()

        f_score = next((s for d, s in FREQ_OPTIONS if d == freq), 0)
        l_score = next((s for d, s in LIKE_OPTIONS if d == like), 0)
        s_score = next((s for d, s in SEV_OPTIONS if d == sev), 0)

        score = RISK_MATRIX.get((l_score, s_score), 0)
        level, color = get_risk_level(score)

        self.risk_score_label.config(text=f"درجة المخاطرة: {score}", fg=color)
        self.risk_level_label.config(text=f"مستوى الخطورة: {level}", fg=color)

    def load_record(self):
        r = self.record
        self.job_title.insert(0, r.job_title)
        self.workplace.insert(0, r.workplace)
        self.date.insert(0, r.date)
        self.analyst.insert(0, r.analyst)
        self.worker.insert(0, r.worker)
        self.general_notes.insert("1.0", r.general_notes)
        if r.department:
            self.dept_form_var.set(r.department)
        self.worker_sig.insert(0, r.worker_sig)
        self.safety_officer.insert(0, r.safety_officer)
        self.dept_manager.insert(0, r.dept_manager)
        self.general_manager.insert(0, r.general_manager)

        if r.freq_desc:
            self.freq_var.set(r.freq_desc)
        if r.likelihood_desc:
            self.like_var.set(r.likelihood_desc)
        if r.severity_desc:
            self.sev_var.set(r.severity_desc)

        for ppe in r.ppe_items:
            if ppe in self.ppe_vars:
                self.ppe_vars[ppe].set(True)

        for i, step in enumerate(r.steps):
            if i < len(self.step_widgets):
                _, ent_step, ent_hazard, ent_control, ent_notes = self.step_widgets[i]
                ent_step.insert("1.0", step.work_step)
                ent_hazard.insert("1.0", step.hazard)
                ent_control.insert("1.0", step.control_method)
                ent_notes.insert("1.0", step.notes)

        self.calc_risk()

    def save(self):
        r = self.record
        r.job_title = self.job_title.get().strip()
        r.workplace = self.workplace.get().strip()
        r.department = self.dept_form_var.get().strip()
        r.date = self.date.get().strip()
        r.analyst = self.analyst.get().strip()
        r.worker = self.worker.get().strip()
        r.general_notes = self.general_notes.get("1.0", "end").strip()
        r.worker_sig = self.worker_sig.get().strip()
        r.safety_officer = self.safety_officer.get().strip()
        r.dept_manager = self.dept_manager.get().strip()
        r.general_manager = self.general_manager.get().strip()

        r.freq_desc = self.freq_var.get()
        r.freq_score = next((s for d, s in FREQ_OPTIONS if d == r.freq_desc), 1)
        r.likelihood_desc = self.like_var.get()
        r.likelihood_score = next((s for d, s in LIKE_OPTIONS if d == r.likelihood_desc), 1)
        r.severity_desc = self.sev_var.get()
        r.severity_score = next((s for d, s in SEV_OPTIONS if d == r.severity_desc), 1)

        r.risk_score = RISK_MATRIX.get((r.likelihood_score, r.severity_score), 0)
        r.risk_level, _ = get_risk_level(r.risk_score)

        r.ppe_items = [ppe for ppe, var in self.ppe_vars.items() if var.get()]

        r.steps = []
        for ent_num, ent_step, ent_hazard, ent_control, ent_notes in self.step_widgets:
            step_text = ent_step.get("1.0", "end").strip()
            hazard_text = ent_hazard.get("1.0", "end").strip()
            control_text = ent_control.get("1.0", "end").strip()
            notes_text = ent_notes.get("1.0", "end").strip()
            if step_text or hazard_text or control_text:
                num = int(ent_num.get().strip() or "0")
                r.steps.append(WorkStep(num, step_text, hazard_text, control_text, notes_text))

        if not r.job_title:
            messagebox.showwarning("تنبيه", "الرجاء إدخال الوظيفة")
            return

        if self.is_new:
            db.add(r)
        else:
            db.update(r)

        self.app.list_view.refresh()
        self.destroy()
        messagebox.showinfo("نجاح", "تم الحفظ بنجاح")

class JSAViewForm(tk.Toplevel):
    def __init__(self, app, record: JSARecord):
        super().__init__(app.root)
        self.title(f"عرض JSA - {record.job_title}")
        self.geometry("1000x750")
        self.configure(bg="white")

        canvas = tk.Canvas(self, bg="white")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        frame = tk.Frame(canvas, bg="white")
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw", width=960)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        header = tk.Frame(frame, bg="#1a237e")
        header.pack(fill="x")
        logo_path = os.path.join(os.path.dirname(__file__), "namaa logo.png")
        for p in [logo_path, logo_path.replace("namaa logo.png", "nama_logo.png")]:
            if os.path.exists(p):
                logo_path = p
                break
        try:
            img = ImageTk.PhotoImage(Image.open(logo_path).resize((70, 70), Image.LANCZOS))
            lbl = tk.Label(header, image=img, bg="#1a237e")
            lbl.image = img
            lbl.pack(side="right", padx=(10, 5), pady=5)
        except:
            pass
        texts = tk.Frame(header, bg="#1a237e")
        texts.pack(side="right", padx=(5, 10))
        tk.Label(texts, text="شركة نماء لصناعة الأعلاف", font=("Arial", 16, "bold"),
                 bg="#1a237e", fg="white").pack(anchor="e")
        tk.Label(texts, text="Job Safety Analysis\nتحليل مخاطر الوظيفة",
                  font=("Arial", 12), bg="#1a237e", fg="#B0BEC5").pack(anchor="e")

        info = tk.Frame(frame, bg="white", padx=20, pady=10)
        info.pack(fill="x")
        fields = [
            ("الوظيفة:", record.job_title),
            ("مكان العمل:", record.workplace),
            ("التاريخ:", record.date),
            ("القائم بالتحليل:", record.analyst),
            ("القائم بالعمل:", record.worker),
        ]
        for i, (lbl, val) in enumerate(fields):
            row_f = tk.Frame(info, bg="white")
            row_f.pack(fill="x", pady=2)
            tk.Label(row_f, text=lbl, font=("Arial", 10, "bold"), bg="white",
                     width=18, anchor="e").pack(side="right")
            tk.Label(row_f, text=val, font=("Arial", 10), bg="#f5f5f5",
                     anchor="w", padx=5).pack(side="right", fill="x", expand=True)

        risk_f = tk.LabelFrame(frame, text="تقييم المخاطر", font=("Arial", 12, "bold"),
                               bg="white", fg="#1a237e", padx=10, pady=10)
        risk_f.pack(fill="x", padx=20, pady=10)

        level, color = get_risk_level(record.risk_score)
        for i, (lbl, val) in enumerate([
            ("التكرار:", record.freq_desc or "--"),
            ("الاحتمالية:", record.likelihood_desc or "--"),
            ("الخطورة:", record.severity_desc or "--"),
            ("درجة المخاطرة:", str(record.risk_score)),
            ("مستوى الخطورة:", f"{level} ({record.risk_level})"),
        ]):
            row_f = tk.Frame(risk_f, bg="white")
            row_f.pack(fill="x", pady=2)
            tk.Label(row_f, text=lbl, font=("Arial", 10, "bold"), bg="white",
                     width=18, anchor="e").pack(side="right")
            tk.Label(row_f, text=val, font=("Arial", 10), bg="#f5f5f5",
                     anchor="w", padx=5).pack(side="right", fill="x", expand=True)

        ppe_f = tk.LabelFrame(frame, text="مهمات الوقاية الشخصية", font=("Arial", 12, "bold"),
                              bg="white", fg="#1a237e", padx=10, pady=10)
        ppe_f.pack(fill="x", padx=20, pady=10)
        ppe_text = " - ".join(record.ppe_items) if record.ppe_items else "لا يوجد"
        tk.Label(ppe_f, text=ppe_text, font=("Arial", 10), bg="white", anchor="w").pack(fill="x")

        steps_f = tk.LabelFrame(frame, text="خطوات العمل وتحليل المخاطر", font=("Arial", 12, "bold"),
                                bg="white", fg="#1a237e", padx=10, pady=10)
        steps_f.pack(fill="x", padx=20, pady=10)

        cols = ["م", "خطوات العمل", "الاخطار المحتمله", "طرق التحكم"]
        header_f = tk.Frame(steps_f, bg="#e8eaf6")
        header_f.pack(fill="x")
        for c in cols:
            tk.Label(header_f, text=c, font=("Arial", 9, "bold"), bg="#e8eaf6",
                     width=8 if c == "م" else 25, anchor="center").pack(side="right", padx=1)

        for step in record.steps:
            row_st = tk.Frame(steps_f, bg="white")
            row_st.pack(fill="x", pady=1)
            tk.Label(row_st, text=str(step.number), font=("Arial", 9), bg="white",
                     width=8, anchor="center").pack(side="right")
            tk.Label(row_st, text=step.work_step, font=("Arial", 9), bg="#fafafa",
                     width=25, anchor="w", wraplength=200).pack(side="right", padx=1)
            tk.Label(row_st, text=step.hazard, font=("Arial", 9), bg="#fafafa",
                     width=25, anchor="w", wraplength=200).pack(side="right", padx=1)
            tk.Label(row_st, text=step.control_method, font=("Arial", 9), bg="#fafafa",
                     width=25, anchor="w", wraplength=200).pack(side="right", padx=1)

        sig_f = tk.LabelFrame(frame, text="التوقيعات", font=("Arial", 12, "bold"),
                              bg="white", fg="#1a237e", padx=10, pady=10)
        sig_f.pack(fill="x", padx=20, pady=10)
        for lbl, val in [
            ("السلامة والصحة المهنية:", record.safety_officer),
            ("القائم بالعملية:", record.worker_sig),
            ("مدير الإدارة:", record.dept_manager),
            ("المدير العام:", record.general_manager),
        ]:
            row_f = tk.Frame(sig_f, bg="white")
            row_f.pack(fill="x", pady=2)
            tk.Label(row_f, text=lbl, font=("Arial", 10, "bold"), bg="white",
                     width=25, anchor="e").pack(side="right")
            tk.Label(row_f, text=val or "--", font=("Arial", 10), bg="#f5f5f5",
                     anchor="w", padx=5).pack(side="right", fill="x", expand=True)

        if record.general_notes:
            note_f = tk.LabelFrame(frame, text="ملاحظات", font=("Arial", 12, "bold"),
                                   bg="white", fg="#1a237e", padx=10, pady=10)
            note_f.pack(fill="x", padx=20, pady=10)
            tk.Label(note_f, text=record.general_notes, font=("Arial", 10),
                     bg="white", anchor="w", wraplength=800).pack(fill="x")

        tk.Button(frame, text="إغلاق", font=("Arial", 11), bg="#f44336", fg="white",
                  padx=30, command=self.destroy).pack(pady=15)

class JSAApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("نظام تحليل مخاطر الوظيفة - JSA")
        self.root.geometry("1100x650")
        self.root.minsize(900, 500)
        self.root.configure(bg="#f0f4f8")

        try:
            self.root.iconbitmap(default=os.path.join(os.path.dirname(__file__), "namaa logo.png"))
        except:
            pass

        self.list_view = JSAListView(self.root, self)
        self.list_view.pack(fill="both", expand=True)

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self):
        db.save()
        self.root.destroy()

    def new_record(self):
        JSAFormView(self)

    def view_record(self):
        rid = self.list_view.get_selected_id()
        if rid is None:
            return
        record = next((r for r in db.records if r.id == rid), None)
        if record:
            JSAViewForm(self, record)

    def edit_record(self):
        rid = self.list_view.get_selected_id()
        if rid is None:
            return
        record = next((r for r in db.records if r.id == rid), None)
        if record:
            JSAFormView(self, record)

    def delete_record(self):
        rid = self.list_view.get_selected_id()
        if rid is None:
            return
        if messagebox.askyesno("تأكيد الحذف", "هل أنت متأكد من حذف هذا السجل؟"):
            db.delete(rid)
            self.list_view.refresh()

    def _refresh_depts(self):
        self.list_view.refresh()
        self.list_view._update_dept_filter()

    def open_dept_mgr(self):
        DeptManager(self)

    def open_analysis(self):
        AnalysisDashboard(self)

    def import_excel(self):
        from tkinter import filedialog as fd
        path = fd.askopenfilename(
            title="اختر ملف إكسيل",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            imported = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                def cv(r, c):
                    v = ws.cell(row=r, column=c).value
                    return str(v).strip() if v else ""

                job_title = cv(5, 4) or cv(5, 2)
                if not job_title or job_title in ("الوظيفة", ""):
                    continue

                workplace = cv(5, 10) or cv(5, 9)
                date = cv(7, 4) or cv(7, 2)
                analyst = cv(7, 10) or cv(7, 8)
                worker = cv(24, 10) or cv(24, 8)
                safety_officer = cv(24, 2)
                dept_manager = cv(24, 14) or cv(24, 12)
                general_manager = cv(24, 19) or cv(24, 17)
                general_notes = cv(18, 3)

                try:
                    freq_score = int(float(cv(9, 11) or "1"))
                except:
                    freq_score = 1
                try:
                    like_score = int(float(cv(11, 11) or "1"))
                except:
                    like_score = 1
                try:
                    sev_score = int(float(cv(13, 11) or "1"))
                except:
                    sev_score = 1

                freq_desc = cv(9, 5) or cv(9, 3)
                like_desc = cv(11, 5) or cv(11, 3)
                sev_desc = cv(13, 5) or cv(13, 3)

                try:
                    risk_score = int(float(cv(15, 11) or cv(15, 10) or "0"))
                except:
                    risk_score = RISK_MATRIX.get((like_score, sev_score), 0)
                risk_level, _ = get_risk_level(risk_score)

                ppe_set = set()
                for r in (9, 11, 13, 15):
                    for c in range(14, 21):
                        v = cv(r, c)
                        if v and v not in ("0", "مهمات الوقاية", "مهمات الوقاية "):
                            ppe_set.add(v)

                steps = []
                for r in (22, 23):
                    try:
                        n = int(float(cv(r, 2)))
                    except:
                        n = r - 21
                    st = cv(r, 3)
                    hz = cv(r, 8)
                    ct = cv(r, 12)
                    nt = cv(r, 17)
                    if st or hz or ct:
                        steps.append(WorkStep(n, st, hz, ct, nt))

                dept = "أخرى"
                for kw, d in {"الشئون الادارية": "الشئون الإدارية",
                               "مكتب": "الشئون الإدارية"}.items():
                    if kw in workplace:
                        dept = d
                        break

                record = JSARecord(
                    job_title=job_title, workplace=workplace, department=dept,
                    date=date, analyst=analyst, freq_desc=freq_desc,
                    freq_score=freq_score, likelihood_desc=like_desc,
                    likelihood_score=like_score, severity_desc=sev_desc,
                    severity_score=sev_score, risk_score=risk_score,
                    risk_level=risk_level, ppe_items=list(ppe_set),
                    steps=steps, worker=worker, safety_officer=safety_officer,
                    dept_manager=dept_manager, general_manager=general_manager,
                    general_notes=general_notes,
                )

                existing = [r for r in db.records if r.job_title == job_title]
                if existing:
                    record.id = existing[0].id
                    db.update(record)
                else:
                    db.add(record)
                imported += 1

            wb.close()
            self.list_view.refresh()
            messagebox.showinfo("نجاح", f"تم استيراد {imported} سجل/سجلات بنجاح")
        except Exception as e:
            import traceback
            messagebox.showerror("خطأ", f"فشل الاستيراد:\n{str(e)}\n\n{traceback.format_exc()}")

    def generate_report(self):
        rid = self.list_view.get_selected_id()
        if rid is None:
            return
        record = next((r for r in db.records if r.id == rid), None)
        if not record:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as pdf_canvas
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import arabic_reshaper
            from bidi.algorithm import get_display

            def ar(text):
                return get_display(arabic_reshaper.reshape(text))

            # Register Arabic TTF fonts
            pdfmetrics.registerFont(TTFont('Ar', r"C:\Windows\Fonts\arial.ttf"))
            pdfmetrics.registerFont(TTFont('ArBd', r"C:\Windows\Fonts\arialbd.ttf"))

            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"JSA_{record.job_title}_{record.date}.pdf"
            )
            if not file_path:
                return

            pdf = pdf_canvas.Canvas(file_path, pagesize=A4)
            W, H = A4
            LM = 30          # left margin
            RM = W - 30      # right margin x
            ENG_X = LM       # x for English labels (drawString LTR)
            VAL_R = 400      # x for values RIGHT edge (drawRightString)
            AR_X = RM        # x for Arabic labels RIGHT edge (drawRightString)
            PAGE_TOP = H - 30

            # ---- helpers ----
            def draw_section_title(y, title):
                """Draw a section header bar and title, returns new y"""
                pdf.setFillColorRGB(0.91, 0.92, 0.96)  # #e8eaf6 light gray-blue
                pdf.rect(LM - 10, y - 2, RM - LM + 20, 22, fill=1, stroke=0)
                pdf.setFillColor(colors.black)
                pdf.setFont("ArBd", 11)
                pdf.drawRightString(AR_X, y + 4, ar(title))
                return y - 28

            def draw_three_col(y, eng_label, value, ar_label, font_size=10):
                """Draw English | value | Arabic in one row, returns new y"""
                pdf.setFont("Ar", font_size)
                pdf.drawString(ENG_X, y, eng_label)
                pdf.drawRightString(AR_X, y, ar(ar_label))
                pdf.drawRightString(VAL_R, y, ar(str(value)))
                return y - 20

            def draw_field_row(y, eng_label, key, ar_label):
                v = getattr(record, key, "")
                return draw_three_col(y, eng_label, v, ar_label)

            def check_page(y, needed=80):
                if y < needed:
                    pdf.showPage()
                    return PAGE_TOP
                return y

            # ================================================================
            #  HEADER BAR (dark blue, matching UI)
            # ================================================================
            pdf.setFillColorRGB(0.10, 0.14, 0.49)  # #1a237e
            pdf.rect(0, H - 95, W, 95, fill=1, stroke=0)
            pdf.setFillColor(colors.white)

            # Logo
            logo_path = os.path.join(os.path.dirname(__file__), "namaa logo.png")
            for alt in [logo_path, logo_path.replace("namaa logo.png", "nama_logo.png")]:
                if os.path.exists(alt):
                    logo_path = alt
                    break
            try:
                from PIL import Image as PILImg
                img = PILImg.open(logo_path)
                lw, lh = img.size
                logo_w = 65
                logo_h = lh * (logo_w / lw)
                pdf.drawImage(logo_path, RM - logo_w - 5, H - 85,
                              width=logo_w, height=logo_h,
                              preserveAspectRatio=True)
            except:
                pass

            pdf.setFont("ArBd", 16)
            pdf.drawRightString(RM - 80, H - 38, ar("شركة نماء لصناعة الأعلاف"))
            pdf.setFont("Ar", 10)
            pdf.setFillColorRGB(0.69, 0.70, 0.77)  # #B0BEC5
            pdf.drawRightString(RM - 80, H - 58, "Job Safety Analysis")
            pdf.setFillColor(colors.white)
            pdf.setFont("ArBd", 13)
            pdf.drawRightString(RM - 80, H - 76, ar("تحليل مخاطر الوظيفة"))

            y = H - 110

            # ================================================================
            #  BASIC INFO -  البيانات الأساسية
            # ================================================================
            y = check_page(y, 200)
            y = draw_section_title(y, "البيانات الأساسية")
            y = draw_field_row(y, "JOB TITLE",     "job_title",     "الوظيفة:")
            y = draw_field_row(y, "WORKPLACE",     "workplace",     "مكان العمل:")
            y = draw_field_row(y, "DEPARTMENT",    "department",    "الإدارة:")
            y = draw_field_row(y, "DATE",          "date",          "التاريخ:")
            y = draw_field_row(y, "ANALYST",       "analyst",       "القائم بالتحليل:")
            y = draw_field_row(y, "WORKER",        "worker",        "القائم بالعمل:")

            # GENERAL NOTES (Text field like the UI)
            y -= 4
            pdf.setFont("ArBd", 10)
            pdf.drawString(ENG_X, y, "GENERAL NOTES")
            pdf.drawRightString(AR_X, y, ar("الملاحظات العامة:"))
            y -= 18
            pdf.setFont("Ar", 9)
            if record.general_notes:
                for line in record.general_notes.split("\n"):
                    y = check_page(y, 50)
                    pdf.drawRightString(VAL_R, y, ar(line.strip()))
                    y -= 14
            else:
                pdf.drawRightString(VAL_R, y, "---")
                y -= 14
            y -= 10

            # ================================================================
            #  RISK ASSESSMENT - تقييم المخاطر
            # ================================================================
            y = check_page(y, 180)
            y = draw_section_title(y, "تقييم المخاطر")
            y = draw_three_col(y, "FREQUENCY",
                               record.freq_desc or "---",
                               "التكرار:")
            y = draw_three_col(y, "LIKELIHOOD",
                               record.likelihood_desc or "---",
                               "الاحتمالية:")
            y = draw_three_col(y, "SEVERITY",
                               record.severity_desc or "---",
                               "الخطورة:")

            y -= 5
            level, _ = get_risk_level(record.risk_score)
            pdf.setFont("ArBd", 11)
            pdf.setFillColorRGB(0.10, 0.14, 0.49)
            pdf.drawRightString(VAL_R, y, ar(f"درجة المخاطرة: {record.risk_score}"))
            pdf.drawRightString(AR_X, y, ar(f"مستوى الخطورة: {level}"))
            pdf.setFillColor(colors.black)
            y -= 22

            # ================================================================
            #  PPE - مهمات الوقاية الشخصية
            # ================================================================
            y -= 5
            y = draw_section_title(y, "مهمات الوقاية الشخصية (PPE)")
            ppe_list = ["النظارات الواقية", "جوانتى", "درع لحام",
                        "ملابس واقية", "أحذية للانزلاق",
                        "أحذية آمنة", "جهاز تنفس", "خوذة",
                        "واقي سمع", "حزام أمان"]
            ppe_cols = 5
            PPE_LEFT = LM + 30
            cell_w = (RM - PPE_LEFT) / ppe_cols
            pdf.setFont("Ar", 9)
            for i, ppe in enumerate(ppe_list):
                checked = "✓" if ppe in record.ppe_items else "☐"
                x = PPE_LEFT + (i % ppe_cols) * cell_w
                pdf.drawString(x, y, f"{checked} {ar(ppe)}")
                if i % ppe_cols == ppe_cols - 1:
                    y -= 16
            if len(ppe_list) % ppe_cols != 0:
                y -= 16
            y -= 10

            # ================================================================
            #  STEPS TABLE - خطوات العمل وتحليل المخاطر
            # ================================================================
            y = check_page(y, 120)
            y = draw_section_title(y, "خطوات العمل وتحليل المخاطر")

            # Table columns (matching UI: الملاحظات | طرق التحكم | الاخطار المحتمله | خطوات العمل | م)
            col_headers = ["الملاحظات", "طرق التحكم", "الاخطار المحتمله", "خطوات العمل", "م"]
            col_widths  = [80,          150,           130,                150,            25]
            col_chars   = [20,          35,            30,                 35,             3]
            # x positions: draw from right to left (RTL)
            txs = [RM]
            for cw in col_widths:
                txs.append(txs[-1] - cw)

            # Header row
            pdf.setFillColorRGB(0.91, 0.92, 0.96)
            pdf.rect(LM - 10, y - 2, RM - LM + 20, 18, fill=1, stroke=0)
            pdf.setFillColor(colors.black)
            pdf.setFont("ArBd", 9)
            for i, (hdr, cw) in enumerate(zip(col_headers, col_widths)):
                pdf.drawCentredString(txs[i] - cw / 2, y + 2, ar(hdr))
            y -= 18

            # Data rows
            pdf.setFont("Ar", 8)
            for step in record.steps:
                if y < 55:
                    pdf.showPage()
                    y = PAGE_TOP
                    # Redraw section title + header
                    y = draw_section_title(y, "خطوات العمل وتحليل المخاطر (تابع)")
                    pdf.setFillColorRGB(0.91, 0.92, 0.96)
                    pdf.rect(LM - 10, y - 2, RM - LM + 20, 18, fill=1, stroke=0)
                    pdf.setFillColor(colors.black)
                    pdf.setFont("ArBd", 9)
                    for i, (hdr, cw) in enumerate(zip(col_headers, col_widths)):
                        pdf.drawCentredString(txs[i] - cw / 2, y + 2, ar(hdr))
                    y -= 18
                    pdf.setFont("Ar", 8)

                row_data = [step.notes, step.control_method, step.hazard,
                            step.work_step, str(step.number)]
                for i, (txt, cw, mc) in enumerate(zip(row_data, col_widths, col_chars)):
                    x_right = txs[i]
                    if i == len(row_data) - 1:  # number column - center
                        pdf.drawCentredString(x_right - cw / 2, y, txt)
                    else:
                        pdf.drawRightString(x_right - 3, y, ar(txt[:mc]))
                y -= 16
            y -= 10

            # ================================================================
            #  SIGNATURES - التوقيعات
            # ================================================================
            y = check_page(y, 120)
            y = draw_section_title(y, "التوقيعات")
            sigs = [
                ("SAFETY OFFICER",   "safety_officer",  "السلامة والصحة المهنية:"),
                ("WORKER",           "worker_sig",      "القائم بالعملية:"),
                ("DEPT. MANAGER",    "dept_manager",    "مدير الإدارة:"),
                ("GENERAL MANAGER",  "general_manager", "المدير العام:"),
            ]
            for eng_label, key, ar_label in sigs:
                v = getattr(record, key, "") or "---"
                pdf.setFont("Ar", 10)
                pdf.drawString(ENG_X, y, eng_label)
                pdf.drawRightString(AR_X, y, ar(ar_label))
                pdf.drawRightString(VAL_R, y, ar(v))
                # Underline for signature
                pdf.setStrokeColorRGB(0.8, 0.8, 0.8)
                pdf.line(ENG_X + 100, y - 2, VAL_R, y - 2)
                pdf.setStrokeColor(colors.black)
                y -= 20

            pdf.save()
            messagebox.showinfo("نجاح", f"تم حفظ التقرير:\n{file_path}")

        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            messagebox.showerror("خطأ", f"فشل في إنشاء التقرير:\n{str(e)}\n\n{tb}")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = JSAApp()
    app.run()
